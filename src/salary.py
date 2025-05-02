import cl, ex
import openpyxl as op
import os

def create_drivers_table(r_wb):

    zwb = op.load_workbook(f'salary-data/{folder_year}/data/zisseki.xlsx')
    zsh = zwb.active
    g_wb = op.load_workbook('data/gather.xlsx')
    g_sh = g_wb.worksheets[0]

    sheets = r_wb.sheetnames

    r = 2

    for driver in d1_dict:

        t_wb = op.load_workbook('data/salary-table2.xlsx',
                                                      data_only=True)
        zt_sh = t_wb.worksheets[0]
        mt_sh = t_wb.worksheets[1]

        sum_goukei, syukin = ex.drivers_cource(zsh, driver)

        zt_sh, g_sh, k_sum, r, kisozikyu = ex.zangyou_table(zt_sh, g_sh, r, driver,
                                    sum_goukei, syukin, d1_dict, folder_year)

        mt_sh, g_sh, r = ex.salary_table(mt_sh, g_sh, sum_goukei, k_sum,
                                                syukin, d1_dict, driver, r, kisozikyu)

        if driver in sheets:
          r_sh = r_wb[driver]
        else:
          r_sh = r_wb.copy_worksheet(r_wb["Sheet1"])
          r_sh.title = driver

        r_sh = ex.create_d_results(r_sh, g_sh, folder_year, driver)


        t_wb_SAVE_PATH = f'salary-data/{folder_year}/{folder_year}-{driver}.xlsx'
        t_wb.save(t_wb_SAVE_PATH)

        g_wb_SAVE_PATH = f'salary-data/{folder_year}/{folder_year}-results.xlsx'
        g_wb.save(g_wb_SAVE_PATH)

        r_wb_SAVE_PATH = f'salary-data/drivers-results.xlsx'
        r_wb.save(r_wb_SAVE_PATH)

def create_folder():

  path = 'salary-data'
  os.makedirs(path, exist_ok=True)
  path_d = f'salary-data/{folder_year}/data'
  os.makedirs(path_d, exist_ok=True)

def create_dict():

  d1wb = op.load_workbook(f'drivers-data/drivers1.xlsx')
  d1sh = d1wb.active
  columuns = list(d1sh.columns)[0]
  i = 1
  d1_dict = {}
  for rows in d1sh.iter_rows(min_row=2, min_col=2):
    for row in rows:
      key = columuns[i].value
      if key not in d1_dict:
        d1_dict[key] = []
      d1_dict[key].append(row.value)
    i += 1

  return d1_dict

d1_dict = create_dict()

csv_year, folder_year = cl.input_years()

create_folder()

if "drivers-results.xlsx" in os.listdir("salary-data"):
  r_wb = op.load_workbook("salary-data/drivers-results.xlsx")
else:
  r_wb = op.load_workbook('data/gather.xlsx')

ex.convert_excel(csv_year, folder_year)

create_drivers_table(r_wb)

cl.create_ok_message()


