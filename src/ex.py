import cl
import pandas as pd
import openpyxl as op
from openpyxl.styles import Alignment
from decimal import *
import zipfile
import os
import math

def convert_excel(csv_year, folder_year):

    #zisseki_path = f'C:/CSVファイル/配車実績ＣＳＶ({csv_year})-({csv_year}).CSV'
    zisseki_path = f'zisseki_csv/配車実績ＣＳＶ({csv_year})-({csv_year}).CSV'
    #file_path_zip = "C:/Users/Mieyuso001/Downloads/労働時間管理表.zip"
    #file_path_zip = "C:/Users/Mieyuso005/Downloads/労働時間管理表.zip"
    file_path_zip = f'kinmuzikan/{folder_year}/労働時間管理表.zip'
    try:
        with zipfile.ZipFile(file_path_zip) as zip_f:
            for info in zip_f.infolist():
                info.filename = info.orig_filename.encode('cp437').decode('cp932')
                if os.sep != '/' and os.sep in info.filename:
                    info.filename = info.filename.replace(os.sep, '/')
                zip_f.extract(info, path=f'salary-data/{folder_year}/data')
    except FileNotFoundError:
        ms = cl.create_error_message\
            ('労働時間管理表.zipを出力しましたか？')

    try:
        data_zisseki = pd.read_csv(zisseki_path, encoding='cp932')
    except:
        ms = cl.create_error_message\
            ('配車実績ＣＳＶを出力しましたか？／年月の入力が間違っていませんか？')
    data_zisseki.to_excel(f'salary-data/{folder_year}/data/zisseki.xlsx',
                                                index=False)
    data_time = pd.read_csv(f'salary-data/{folder_year}/data/労働時間管理表(合計)_00000001.csv',
                                            encoding='cp932')
    data_time.to_excel(f'salary-data/{folder_year}/data/time.xlsx', index=False)

def drivers_cource(zsh, driver):

    sum_goukei = 0
    syukin_list = []
    for row in range(2, zsh.max_row+1):
        if driver == zsh.cell(row, 9).value:
            if zsh.cell(row, 1).value not in syukin_list:
                syukin_list.append(str(zsh.cell(row, 1).value))
            sum_goukei += int((zsh.cell(row, 27).value).replace('\\', ''))
    syukin = len(syukin_list)

    return sum_goukei, syukin

def return_HM(zikan):

    time_tap = math.modf(zikan)
    f = time_tap[0]
    t = int(time_tap[1])
    m = round(Decimal(str(f * 60)), 0)
    m = '{:0>2}'.format(m)
    HM = '{0}:{1}'.format(t, m)

    return HM

def convert_zikan(tsh, row, col):

    time = tsh.cell(row, col).value.split(':')
    htime = int(time[0])
    mtime = int(time[1])
    syousuu = Decimal(str(mtime/60))
    zikan = htime + syousuu

    return zikan

def arrange_sheet(sh, r):

    for i in range(3, 14, 2):
        sh.cell(r, i).number_format = '#,###'
    sh.cell(r, 4).number_format = '#,###'
    sh.cell(r, 14).number_format = '#,###'
    for n in range(6, 13, 2):
        sh.cell(r, n).alignment = Alignment(horizontal='right')

def zangyou_table(zt_sh, g_sh, r, driver, buai, syukin, d1_dict, folder_year):

    twb = op.load_workbook(f'salary-data/{folder_year}/data/time.xlsx')
    tsh = twb.worksheets[0]
    for col in range(1, tsh.max_column+1):
        if tsh.cell(1, col).value == "総労働時間":
            sou = col
        if tsh.cell(1, col).value == "休日労働時間":
            kyu = col
        if tsh.cell(1, col).value == "法定外労働時間":
            zan = col
        if tsh.cell(1, col).value == "深夜労働時間":
            shin = col
    kisozikyu = 0
    k_sum= 0

    for row in range(2, tsh.max_row+1):
        if tsh.cell(row, 5).value == driver:
            k_sum = 0
            buai = buai
            souroudou = convert_zikan(tsh, row, sou)
            kyuuzitu = convert_zikan(tsh, row, kyu)
            zangyou = convert_zikan(tsh, row, zan)
            #zangyou = zangyou - kyuuzitu
            shinya =  convert_zikan(tsh, row, shin)
            cyouzan = zangyou - 60
            teate = d1_dict[driver][0] + d1_dict[driver][1]
            kisoteate = Decimal(str(teate / 174))
            try:
                kisozikyu = math.ceil(Decimal(str(buai / souroudou)))
            except:
                kisozikyu = 0
            k_zangyou = kisozikyu * Decimal("0.25")
            if zangyou > 60:
                k_cyouzan = kisozikyu * Decimal("0.5")
            if kyuuzitu != 0:
                k_kyuuzitu = kisozikyu * Decimal("0.35")
            if shinya != 0:
                k_shinya = kisozikyu * Decimal("0.25")
            zt_sh.cell(1, 1).value = driver
            if cyouzan > 0:
                zt_sh.cell(2, 2).value = '60:00'
                zt_sh.cell(2, 4).value = math.ceil(k_zangyou * 60)
                zt_sh.cell(2, 5).value = math.ceil(kisoteate * Decimal("1.25") * 60)
                zt_sh.cell(2, 3).value = zt_sh.cell(2, 4).value + zt_sh.cell(2, 5).value
                k_sum += zt_sh.cell(2, 3).value
                zt_sh.cell(3, 2).value = return_HM(cyouzan)
                zt_sh.cell(3, 4).value = math.ceil(k_cyouzan * cyouzan)
                zt_sh.cell(3, 5).value = math.ceil(kisoteate * Decimal("1.5") * cyouzan)
                zt_sh.cell(3, 3).value = zt_sh.cell(3, 4).value + zt_sh.cell(3, 5).value
                k_sum += zt_sh.cell(3, 3).value
            else:
                zt_sh.cell(2, 2).value = return_HM(zangyou)
                zt_sh.cell(2, 4).value = math.ceil(k_zangyou * zangyou)
                zt_sh.cell(2, 5).value = math.ceil(kisoteate * Decimal("1.25") * zangyou)
                zt_sh.cell(2, 3).value = zt_sh.cell(2, 4).value + zt_sh.cell(2, 5).value
                k_sum += zt_sh.cell(2, 3).value
                zt_sh.cell(3, 3).value = 0
            if kyuuzitu != 0:
                zt_sh.cell(4, 2).value = tsh.cell(row, 14).value
                zt_sh.cell(4, 4).value = math.ceil(k_kyuuzitu * kyuuzitu)
                zt_sh.cell(4, 5).value = math.ceil(kisoteate * Decimal("1.35") * kyuuzitu)
                zt_sh.cell(4, 3).value = zt_sh.cell(4, 4).value + zt_sh.cell(4, 5).value
                k_sum += zt_sh.cell(4, 3).value
            else:
                zt_sh.cell(4, 3).value = 0
            if shinya != 0:
                zt_sh.cell(5, 2).value = tsh.cell(row, 15).value
                zt_sh.cell(5, 4).value = math.ceil(k_shinya * shinya)
                zt_sh.cell(5, 5).value = math.ceil(kisoteate * Decimal("1.25") * shinya)
                zt_sh.cell(5, 3).value = zt_sh.cell(5, 4).value + zt_sh.cell(5, 5).value
                k_sum += zt_sh.cell(5, 3).value
            else:
                zt_sh.cell(5, 3).value = 0
            zt_sh.cell(6, 3).value = k_sum

    g_sh.cell(r, 1).value = driver
    g_sh.cell(r, 2).value = syukin
    g_sh.cell(r, 3).value = kisozikyu
    g_sh.cell(r, 3).number_format = '#,###'
    g_sh.cell(r, 5).value = buai
    g_sh.cell(r, 5).number_format = '#,###'
    g_sh.cell(r, 6).value = zt_sh.cell(2, 2).value
    g_sh.cell(r, 6).alignment = Alignment(horizontal='right')
    g_sh.cell(r, 7).value = zt_sh.cell(2, 3).value
    g_sh.cell(r, 7).number_format = '#,###'
    g_sh.cell(r, 8).value = zt_sh.cell(3, 2).value
    g_sh.cell(r, 8).alignment = Alignment(horizontal='right')
    g_sh.cell(r, 9).value = zt_sh.cell(3, 3).value
    g_sh.cell(r, 9).number_format = '#,###'
    g_sh.cell(r, 10).value = zt_sh.cell(4, 2).value
    g_sh.cell(r, 10).alignment = Alignment(horizontal='right')
    g_sh.cell(r, 11).value = zt_sh.cell(4, 3).value
    g_sh.cell(r, 11).number_format = '#,###'
    g_sh.cell(r, 12).value = zt_sh.cell(5, 2).value
    g_sh.cell(r, 12).alignment = Alignment(horizontal='right')
    g_sh.cell(r, 13).value = zt_sh.cell(5, 3).value
    g_sh.cell(r, 13).number_format = '#,###'
    g_sh.cell(r, 14).value = k_sum
    g_sh.cell(r, 14).number_format = '#,###'

    return zt_sh, g_sh, k_sum, r, kisozikyu

def salary_table(mt_sh, g_sh, sum_goukei, k_sum, syukin,
                                        d1_dict, driver, r, kisozikyu):

    kyuryou = 0

    mt_sh.cell(1, 1).value = driver
    mt_sh.cell(1, 2).value = round(kisozikyu, 0)
    mt_sh.cell(1, 2).number_format = '#,###'
    mt_sh.cell(2, 2).value = syukin
    mt_sh.cell(3, 2).value = sum_goukei
    mt_sh.cell(4, 2).value = k_sum

    if k_sum != 0:
        mt_sh.cell(5, 2).value = d1_dict[driver][0]
        mt_sh.cell(6, 2).value = d1_dict[driver][1]
        mt_sh.cell(7, 2).value = d1_dict[driver][2]

        for n in range(3, 11):
            if mt_sh.cell(n, 2).value is None:
                pass
            else:
                meisai = int(mt_sh.cell(n, 2).value)
                kyuryou += meisai
    else:
        mt_sh.cell(5, 2).value = 0
        mt_sh.cell(6, 2).value = 0
        mt_sh.cell(7, 2).value = 0
        kyuryou = 0

    mt_sh.cell(8, 2).value = "=B1*C8*8"
    mt_sh.cell(11, 2).value = "=sum(B3:B10)"
    g_sh.cell(r, 4).value = kyuryou
    g_sh.cell(r, 4).number_format = '#,###'
    r += 1

    return mt_sh, g_sh, r

def find_re_sh_row(g_sh, driver):

    row = 2
    while True:

        if driver == g_sh.cell(row, 1).value:
            break
        else:
            row += 1

    return row

def create_d_results(r_sh, g_sh, folder_year, driver):

    row = find_re_sh_row(g_sh, driver)

    r=2
    while True:

        if r_sh.cell(r, 1).value == None:
            r_sh.cell(r, 1).value = folder_year
            for col in range(2, 15):
                r_sh.cell(r, col).value = g_sh.cell(row, col).value
            break
        else:
            r += 1

    r_sh = arrange_sheet(r_sh, r)

    return r_sh





