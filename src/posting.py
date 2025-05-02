import pandas as pd
import openpyxl as op
import cl, math
import datetime
from decimal import *
from dateutil.relativedelta import relativedelta


today = datetime.datetime.now()
now_day = today.strftime('%Y/%m')
b_lastmonth = today + relativedelta(months=-1, day=1)
e_lastmonth = today + relativedelta(day=1, days=-1)
b_last = b_lastmonth.strftime('%m/%d')
e_last = e_lastmonth.strftime('%m/%d')

csv_year, folder_year = cl.input_years()

df = pd.read_csv(f'salary-data/{folder_year}/data/労働時間管理表(合計)_00000001.csv',
                                            encoding='cp932')
ti_df = df[["乗務員名", "総労働時間", '法定外労働時間', '休日労働時間', '深夜労働時間']].set_index("乗務員名")
ti_df = ti_df.T
ti_dict = ti_df.to_dict(orient='list')

so_path = "data/賃金計算連絡書.xlsx"
wb = op.load_workbook(so_path)
sh = wb.active

sh.cell(1, 6).value = now_day
sh.cell(3, 29).value = f"{now_day}/15日支払"
sh.cell(2, 5).value = f"{b_last} ～ {e_last}"

chi_dict ={"役職手当": 6, "通勤手当": 7, "有休支給額": 8, "割増手当":9,
            "職務手当": 11, "歩合給": 12, "無事故手当": 13, "特別手当": 14,
            "出勤日数": 31,}

def convert_zikan(zikan):

    time = zikan.split(':')
    htime = int(time[0])
    mtime = int(time[1])
    syousuu = Decimal(str(mtime/60))
    zikan = htime + syousuu

    return zikan

def return_HM(zikan):

    time_tap = math.modf(zikan)
    f = time_tap[0]
    t = int(time_tap[1])
    m = round(Decimal(str(f * 60)), 0)
    m = '{:0>2}'.format(m)
    HM = '{0}:{1}'.format(t, m)

    return HM

for driver in ti_dict:
    sa_path = f'drivers-data/{folder_year}/{folder_year}-{driver}.xlsx'
    try:
        sa_df = pd.read_excel(sa_path, sheet_name='給与明細', index_col=0)
    except FileNotFoundError:
        continue
    sa_df = sa_df.T
    sa_dict = sa_df.to_dict(orient='list')

    for row in range(15, sh.max_row+1):
        if sh.cell(row, 3).value == driver:
            try:
                sh.cell(row, 32).value = sa_dict["有休支給額"][1]
            except IndexError:
                pass
            for key, col in chi_dict.items():
                for k, v in sa_dict.items():
                    if key == k:
                        sh.cell(row, col).value = sa_dict[k][0]
            sh.cell(row, 33).value = ti_dict[driver][0]
            zan_zikan = convert_zikan(ti_dict[driver][1])
            if ti_dict[driver][2] == "0:00":
                pass
            else:
                kyuzitu = convert_zikan(ti_dict[driver][2])
                #
                # 
                # zan_zikan = zan_zikan - kyuzitu
            if zan_zikan <= 60:
                sh.cell(row, 34).value = ti_dict[driver][1]
            else:
                c_zikan = zan_zikan-60
                c_HM = return_HM(c_zikan)
                sh.cell(row, 34).value = "60:00"
                sh.cell(row, 35).value = c_HM
            if ti_dict[driver][2] != "0:00":
                sh.cell(row, 36).value = ti_dict[driver][2]
            if ti_dict[driver][3] != "0:00":
                sh.cell(row, 37).value = ti_dict[driver][3]

wb.save(f"{folder_year}月末締 賃金計算連絡書 .xlsx")

